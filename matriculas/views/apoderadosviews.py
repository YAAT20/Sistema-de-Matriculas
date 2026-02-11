from matriculas.models import *
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import *
from matriculas.forms import *
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponseRedirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.http import require_POST
import urllib.parse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from django.db.models import Q
from django.http import HttpResponse

from matriculas.views.admin import es_admin_check, SoloAdminMixin

# Vistas para Apoderados
class ApoderadoListView(LoginRequiredMixin, ListView):
    model = Apoderado
    template_name = 'matriculas/apoderados/apoderado_list.html'
    context_object_name = 'apoderados'
    paginate_by = 200

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                models.Q(nombre_completo__icontains=search_query) |
                models.Q(dni__icontains=search_query) |
                models.Q(codigo__icontains=search_query) |
                models.Q(celular=search_query)
            )
        return queryset.order_by('codigo')

class ApoderadoCreateView(LoginRequiredMixin, CreateView):
    model = Apoderado
    form_class = ApoderadoForm
    template_name = 'matriculas/apoderados/apoderado_form.html'
    success_url = reverse_lazy('matriculas:apoderado_list')

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        
        alumno_id = self.request.GET.get('alumno_id')
        if alumno_id:
            try:
                alumno = Alumno.objects.get(pk=alumno_id)
                # Incluir al alumno aún si ya tiene apoderado
                form.fields['alumnos'].queryset = Alumno.objects.filter(
                    models.Q(apoderados__isnull=True) | models.Q(pk=alumno.pk)
                ).distinct()
                form.fields['alumnos'].initial = [alumno]
            except Alumno.DoesNotExist:
                pass
        return form

    def form_valid(self, form):
        apoderado = form.save(commit=False)
        alumnos = form.cleaned_data.get('alumnos')

        if not alumnos:
            form.add_error('alumnos', 'Debe seleccionar al menos un alumno')
            return self.form_invalid(form)

        # Generar código con el primer alumno
        primer_alumno = alumnos.first()
        apoderado.codigo = CodigoManager.generar_codigo_apoderado(primer_alumno.codigo)
        apoderado.save()
        form.save_m2m()  # guarda relación many-to-many
        messages.success(self.request, 'Apoderado registrado exitosamente')
        return redirect(self.success_url)

class ApoderadoUpdateView(LoginRequiredMixin, SoloAdminMixin, UpdateView):
    model = Apoderado
    form_class = ApoderadoForm
    template_name = 'matriculas/apoderados/apoderado_form.html'
    success_url = reverse_lazy('matriculas:apoderado_list')
    #paara asegurarnos de que el user sea admin y esté logeado
    def test_func(self):
        return hasattr(self.request.user, 'perfil') and self.request.user.perfil.tipo == 'admin'
    def form_valid(self, form):
        messages.success(self.request, 'Datos del apoderado actualizados exitosamente')
        return super().form_valid(form)

class ApoderadoDetailView(LoginRequiredMixin, DetailView):
    model = Apoderado
    template_name = 'matriculas/apoderados/apoderado_detail.html'
    context_object_name = 'apoderado'

class ApoderadoDeleteView(LoginRequiredMixin, SoloAdminMixin, DeleteView):
    model = Apoderado
        #paara asegurarnos de que el user sea admin y esté logeado
    def test_func(self):
        return hasattr(self.request.user, 'perfil') and self.request.user.perfil.tipo == 'admin'
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        self.object.delete()
        return JsonResponse({'success': True})

# Vista para asignar apoderado a alumno
class AsignarCrearApoderadoView(LoginRequiredMixin, FormView):
    template_name = 'matriculas/apoderados/asignar_apoderado.html'
    form_class = ApoderadoAlumnoForm

    def dispatch(self, request, *args, **kwargs):
        self.alumno = get_object_or_404(Alumno, pk=self.kwargs.get('alumno_id'))
        return super().dispatch(request, *args, **kwargs)

    def get_initial(self):
        initial = super().get_initial()
        initial['alumno'] = self.alumno
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        search = self.request.GET.get('buscar')
        if search:
            context['resultados'] = Apoderado.objects.filter(dni__icontains=search) | Apoderado.objects.filter(nombre_completo__icontains=search)
        context['alumno'] = self.alumno
        return context

    def form_valid(self, form):
        apoderado = form.save(commit=False)

        # Generar código basado en el alumno si no hay
        if not apoderado.codigo:
            apoderado.codigo = CodigoManager.generar_codigo_apoderado(self.alumno.codigo)

        apoderado.save()  # Guardar primero para que tenga ID

        apoderado.alumnos.add(self.alumno)  # Luego ya puedes usar la relación M2M
        messages.success(self.request, f'Se ha creado y asignado el apoderado {apoderado.nombre_completo}')
        return redirect('matriculas:alumno_detail', pk=self.alumno.pk)
    
    def save(self, *args, **kwargs):
        if not self.pk and hasattr(self, 'alumnos'):
            raise ValueError("No se puede asignar alumnos hasta que el apoderado esté guardado")
        super().save(*args, **kwargs)

@require_POST
@login_required
def asignar_apoderado_existente(request):
    alumno_id = request.POST.get('alumno_id')
    apoderado_id = request.POST.get('apoderado_id')
    alumno = get_object_or_404(Alumno, pk=alumno_id)
    apoderado = get_object_or_404(Apoderado, pk=apoderado_id)
    apoderado.alumnos.add(alumno)
    messages.success(request, f'{apoderado.nombre_completo} asignado a {alumno.nombres_completos}')
    return redirect('matriculas:alumno_detail', pk=alumno.pk)

@login_required
def enviar_ficha_matricula_whatsapp(request, matricula_id):
    matricula = get_object_or_404(Matricula, id=matricula_id)
    apoderado = matricula.apoderado
    if not apoderado:
        messages.error(request, "La matrícula no tiene apoderado asignado.")
        return redirect('matriculas:matricula_detail', pk=matricula_id)

    pdf_url = request.build_absolute_uri(
        reverse('matriculas:ficha_matricula_pdf', args=[matricula.uuid])
    )

    mensaje = (
        f"*¡Bienvenido a la Academia Robert Hooke!* 🎓\n\n"
        f"{apoderado.abreviatura}, *{apoderado.nombre_completo}*, le damos la bienvenida. "
        f"Su {matricula.alumno.sexo_data} *{matricula.alumno.nombres_completos}* ya está matriculado en nuestra academia.\n\n"
        f"📄 *Puede descargar su ficha de matrícula aquí:*\n"
        f"{pdf_url}\n\n"
        f"🔗 *Para acceder al sistema de reportes de avances, simulacros y asistencias:*\n"
        f"https://tinyurl.com/2ygs2dc7 \n\n"
        f"*Sus credenciales son:*\n"
        f"👤 Usuario: {matricula.alumno.codigo}\n"
        f"🔒 Contraseña: {matricula.alumno.dni}\n\n"
        f"⚠️ *IMPORTANTE:* Es obligatorio ingresar al sistema apenas reciba este mensaje o al momento de matricularse, y actualizar sus datos personales.\n"
        f"_(Esta actualización se realiza solo la primera vez que accede)._\n\n"
        f"¡Gracias por confiar en nosotros! 💙\n"
        f"*Academia Robert Hooke*"
    )

    telefono_raw = apoderado.celular if hasattr(apoderado, 'celular') and apoderado.celular else ''
    telefono = ''.join(filter(str.isdigit, str(telefono_raw)))

    if telefono and not telefono.startswith('51'):
        telefono = f'51{telefono}'    
    
    wa_url = f"https://api.whatsapp.com/send?phone={telefono}&text={urllib.parse.quote(mensaje)}"
    
    return HttpResponseRedirect(wa_url)

@login_required
@user_passes_test(es_admin_check)   
def exportar_apoderados_alumnos_excel(request):
    apoderados = Apoderado.objects.all().prefetch_related('alumnos').distinct().order_by('codigo')

    search_query = request.GET.get('search')
    grado_filtro = request.GET.get('grado')

    if search_query:
        apoderados = apoderados.filter(
            Q(nombre_completo__icontains=search_query) |
            Q(dni__icontains=search_query) |
            Q(codigo__icontains=search_query) |
            Q(alumnos__nombres_completos__icontains=search_query) |
            Q(alumnos__codigo__icontains=search_query)
        )

    if grado_filtro:
        apoderados = apoderados.filter(alumnos__grado_estudios=grado_filtro)

    wb = Workbook()
    ws = wb.active
    ws.title = "Apoderados y Alumnos"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") 
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))

    headers = [
        'Cód. Apoderado', 'Nombre Apoderado', 'Celular Apoderado',
        'Cód. Alumno', 'Nombre Alumno', 'Celular Alumno'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    row_num = 2
    
    for apoderado in apoderados:
        sus_alumnos = apoderado.alumnos.all()

        if grado_filtro:
            sus_alumnos = [a for a in sus_alumnos if str(a.grado_estudios) == str(grado_filtro)]

        if sus_alumnos:
            for alumno in sus_alumnos:
                row_data = [
                    apoderado.codigo,
                    apoderado.nombre_completo,
                    apoderado.celular,
                    alumno.codigo,
                    alumno.nombres_completos,
                    alumno.numero_whatsapp
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                row_num += 1
        
        else:
            if not grado_filtro:
                row_data = [
                    apoderado.codigo,
                    apoderado.nombre_completo,
                    apoderado.celular,
                    '-', '-', '-'
                ]
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                row_num += 1

    column_widths = [15, 40, 15, 15, 40, 15]
    for col_num, width in enumerate(column_widths, 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[column_letter].width = width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="Listado_Apoderados_Alumnos_{timestamp}.xlsx"'
    
    wb.save(response)
    return response