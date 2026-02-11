# Modelos
from matriculas import models
from matriculas.models import Alumno, Matricula  

# Django auth y views
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required, user_passes_test

# Views genéricas
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView 
from django.views import View
# Forms
from matriculas.forms import AlumnoForm, MatriculaForm 
#son utilities
from django.contrib import messages
from django.urls import reverse_lazy
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
from django.shortcuts import redirect, get_object_or_404, render
from django.db.models import Q  
from django.urls import reverse

#libs externas
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

from matriculas.views.admin import es_admin_check, SoloAdminMixin

##CRUD ALUMNOS

#LISTA SOLO DE ALUMNOS CON MATRÍCULA ACTIVA
class AlumnoListView(LoginRequiredMixin, ListView):
    model = Alumno
    template_name = 'matriculas/alumno/alumno_list.html'
    context_object_name = 'alumnos'
    paginate_by = 200

    def get_queryset(self):
        queryset = super().get_queryset()

        queryset = queryset.filter(matriculas__estado='activa').distinct()

        search_query = self.request.GET.get('search')
        grado_filtro = self.request.GET.get('grado')

        if search_query:
            queryset = queryset.filter(
                Q(nombres_completos__icontains=search_query) |
                Q(dni__icontains=search_query) |
                Q(codigo__icontains=search_query) |
                Q(numero_whatsapp=search_query)
            )

        if grado_filtro:
            queryset = queryset.filter(grado_estudios=grado_filtro)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['grado_opciones'] = Alumno.GRADO_OPCIONES
        return context

class AlumnoTotalListView(LoginRequiredMixin, ListView):
    model = Alumno
    template_name = 'matriculas/alumno/alumno_total_list.html'
    context_object_name = 'alumnos'
    paginate_by = 40

    def get_queryset(self):
        queryset = super().get_queryset()
        search_query = self.request.GET.get('search')
        grado_filtro = self.request.GET.get('grado')
        estado_filtro = self.request.GET.get('estado')

        # Filtro por búsqueda
        if search_query:
            queryset = queryset.filter(
                Q(nombres_completos__icontains=search_query) |
                Q(dni__icontains=search_query) |
                Q(codigo__icontains=search_query)
            )

        # Filtro por grado
        if grado_filtro:
            queryset = queryset.filter(grado_estudios=grado_filtro)

        # Filtro por estado
        if estado_filtro == 'activo':
            queryset = queryset.filter(activo=True)
        elif estado_filtro == 'inactivo':
            queryset = queryset.filter(activo=False)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['grado_opciones'] = Alumno.GRADO_OPCIONES
        context['estado_actual'] = self.request.GET.get('estado', '')
        context['grado_actual'] = self.request.GET.get('grado', '')

        # 👇 Total real de alumnos (sin depender del paginador)
        context['total_alumnos'] = self.get_queryset().count()
        return context

class AlumnoCreateView(LoginRequiredMixin, CreateView):
    model = Alumno
    form_class = AlumnoForm
    template_name = 'matriculas/alumno/alumno_form.html'
    success_url = reverse_lazy('matriculas:alumno_list')

    def form_valid(self, form):
        messages.success(self.request, 'Alumno registrado exitosamente')
        return super().form_valid(form)

class AlumnoUpdateView(LoginRequiredMixin, SoloAdminMixin, UpdateView):
    model = Alumno
    form_class = AlumnoForm
    template_name = 'matriculas/alumno/alumno_form.html'
    success_url = reverse_lazy('matriculas:alumno_list')

    def form_valid(self, form):
        original = self.get_object()
        if original.grado_estudios != form.cleaned_data['grado_estudios']:
            messages.warning(self.request, "El código del alumno ha sido actualizado por el cambio de grado.")
        messages.success(self.request, "Datos del alumno actualizados exitosamente.")
        return super().form_valid(form)

class AlumnoDetailView(LoginRequiredMixin, DetailView):
    model = Alumno
    template_name = 'matriculas/alumno/alumno_detail.html'
    context_object_name = 'alumno'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        alumno = self.get_object()
        context['tiene_matricula'] = alumno.matriculas.exists()
        context['matricula'] = alumno.matriculas.first()
        return context

class AlumnoDeleteView(LoginRequiredMixin, SoloAdminMixin, DeleteView):
    def post(self, request, pk):
        try:
            alumno = Alumno.objects.get(pk=pk)
            alumno.delete()
            return JsonResponse({'success': True})
        except Alumno.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Alumno no encontrado'}, status=404)
        
@login_required
def enviar_ficha_matricula_whatsapp_estudiante(request, matricula_id):
    matricula = get_object_or_404(Matricula, id=matricula_id)
    alumno = matricula.alumno
    if not alumno:
        messages.error(request, "La matrícula no tiene Alumnos asignado.")
        return redirect('matriculas:matricula_detail', pk=matricula_id)

    pdf_url = request.build_absolute_uri(
        reverse('matriculas:ficha_matricula_pdf', args=[matricula.uuid])
    )

    mensaje = (
        f"*¡Bienvenido a la Academia Robert Hooke!* 🎓\n\n"
        f"Hola, *{alumno.nombres_completos}*, te damos la bienvenida a la academia RH.\n\n"
        f"📄 *Puedes descargar tu ficha de matrícula aquí:*\n"
        f"{pdf_url}\n\n"
        f"🔗 *Para acceder al sistema de reportes de avances, simulacros y asistencias aquí:*\n"
        f"https://tinyurl.com/2ygs2dc7 \n\n"
        f"*Tus credenciales son:*\n"
        f"👤 Usuario: {matricula.alumno.codigo}\n"
        f"🔒 Contraseña: {matricula.alumno.dni}\n\n"
        f"⚠️ *IMPORTANTE:* Es obligatorio ingresar al sistema apenas recibas este mensaje o al momento de matricularte y actualiza tus datos personales.\n"
        f"_(Esta actualización se realiza solo la primera vez que accedes)._\n\n"
        f"¡Gracias por confiar en nosotros! 💙\n"
        f"*Academia Robert Hooke*"
    )

    telefono_raw = alumno.numero_whatsapp if hasattr(alumno, 'numero_whatsapp') and alumno.numero_whatsapp else ''
    telefono = ''.join(filter(str.isdigit, str(telefono_raw)))

    if telefono and not telefono.startswith('51'):
        telefono = f'51{telefono}'

    wa_url = f"https://api.whatsapp.com/send?phone={telefono}&text={urllib.parse.quote(mensaje)}"
    
    return HttpResponseRedirect(wa_url)

@login_required
@user_passes_test(es_admin_check)
def exportar_alumnos_excel(request):
    
    alumnos = Alumno.objects.filter(matriculas__estado='activa').distinct().order_by('codigo')
    
    search_query = request.GET.get('search')
    grado_filtro = request.GET.get('grado')
    
    if search_query:
        alumnos = alumnos.filter(
            Q(nombres_completos__icontains=search_query) |
            Q(dni__icontains=search_query) |
            Q(codigo__icontains=search_query)
        )
    
    if grado_filtro:
        alumnos = alumnos.filter(grado_estudios=grado_filtro)
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Alumnos"
    
    # Definir estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = [
        'nombre',
        'usuario',
        'contraseña'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    for row_num, alumno in enumerate(alumnos, 2):
        row_data = [
            alumno.nombres_completos,
            alumno.codigo,
            alumno.dni,
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    column_widths = [35, 15, 15]
    for col_num, width in enumerate(column_widths, 1):
        column_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[column_letter].width = width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    response['Content-Disposition'] = f'attachment; filename="alumnos_matriculados_{timestamp}.xlsx"'
    
    wb.save(response)
    return response
